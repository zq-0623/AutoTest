from openpyxl.styles import PatternFill
from openpyxl.styles import colors,Font
import openpyxl as vb

#��ȡ ��Ҫ�Աȵ�excel��
workbook_a = vb.load_workbook(r'��1.xlsx')
workbook_b = vb.load_workbook(r'��2.xlsx')
#��ȡ��Ҫ�Աȵ�sheet��
sheet_a = workbook_a['Sheet1']
sheet_b = workbook_b['Sheet1']
#�������е�������
maxrow = sheet_a.max_row
maxcolumn = sheet_b.max_column

#ѭ���Աȱ������е�Ԫ������
for i in range(1,maxrow):
    for j in range(1,maxcolumn):
        cell_a = sheet_a.cell(i,j)
        cell_b = sheet_b.cell(i,j)
        #����в������ݣ��ͱ�ʶ����(��ɫ�Ӵ����壬��ɫ���)��
        if cell_a.value != cell_b.value:
            cell_a.fill = PatternFill("solid",fgColor='FFFF00')
            cell_a.font = Font(color=colors.BLUE,bold=True)
            cell_b.fill = PatternFill("solid",fgColor='FFFF00')
            cell_b.font = Font(color=colors.BLUE,bold=True)
#�����������µ�excel����
workbook_a.save('��1_������.xlsx')
workbook_b.save('��2_������.xlsx')

print("ִ�жԱ����")
